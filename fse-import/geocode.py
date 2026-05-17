#!/usr/bin/env python3
"""
Fill latitude/longitude into fse-research.csv.

For every research-cache row that has no lat/lng yet, look up the
establishment's street address in FSE-review.xlsx (by permit #) and
geocode it with OpenStreetMap Nominatim — the same map data the live
abiteofnutmeg.com Leaflet maps use.

Nominatim asks for <=1 request/sec and a real User-Agent; both are honored.

Run from anywhere:  python3 fse-import/geocode.py
"""

import csv
import json
import os
import re
import time
import urllib.parse
import urllib.request

HERE = os.path.dirname(os.path.abspath(__file__))
REVIEW = os.path.join(HERE, "FSE-review.xlsx")
RESEARCH = os.path.join(HERE, "fse-research.csv")

NOMINATIM = "https://nominatim.openstreetmap.org/search"
HEADERS = {"User-Agent": "abiteofnutmeg-fse-import/1.0 (chrishauman@gmail.com)"}


def clean(text):
    return " ".join(str(text or "").split())


def address_map():
    """permit (str) -> cleaned address string, read from FSE-review.xlsx."""
    import openpyxl
    wb = openpyxl.load_workbook(REVIEW, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    # locate the main table header (col A = "Town", col E = "Facility Name")
    start = next(i for i, r in enumerate(rows)
                 if r[0] == "Town" and r[4] == "Facility Name") + 1
    out = {}
    for r in rows[start:]:
        permit, addr = r[6], r[5]
        out[clean(permit)] = clean(addr)
    return out


def tidy(addr):
    """Turn a messy FSE address into something Nominatim can resolve."""
    a = clean(addr)
    # drop parentheticals like "(town)", "(Stonington)", unit clutter
    a = re.sub(r"\([^)]*\)", "", a)
    a = re.sub(r"\b(ste|suite|unit|apt|bld|bldg)\b\.?\s*\S*", "", a, flags=re.I)
    a = " ".join(a.split())
    if "olde mistick village" in a.lower():
        a = "Olde Mistick Village, Mystic, CT"
    if not re.search(r"\bct\b", a, re.I):
        a += ", CT"
    return a


def geocode(query):
    url = NOMINATIM + "?" + urllib.parse.urlencode(
        {"q": query, "format": "json", "limit": 1, "countrycodes": "us"})
    req = urllib.request.Request(url, headers=HEADERS)
    with urllib.request.urlopen(req, timeout=20) as resp:
        data = json.load(resp)
    if data:
        return round(float(data[0]["lat"]), 6), round(float(data[0]["lon"]), 6)
    return None, None


def main():
    addrs = address_map()
    with open(RESEARCH, newline="", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        fields = reader.fieldnames
        rows = list(reader)

    todo = [r for r in rows if not clean(r.get("lat"))]
    print(f"{len(rows)} cache rows, {len(todo)} need coordinates.")

    done = fail = 0
    for r in todo:
        permit = clean(r["permit"])
        raw = addrs.get(permit, "")
        query = tidy(raw)
        try:
            lat, lng = geocode(query)
        except Exception as e:                       # noqa: BLE001
            lat, lng = None, None
            print(f"  ! {r['name']}: {e}")
        if lat:
            r["lat"], r["lng"] = lat, lng
            done += 1
            print(f"  ok  {r['name']:<34} {lat}, {lng}")
        else:
            fail += 1
            print(f"  --  {r['name']:<34} no match  [{query}]")
        time.sleep(1.1)                              # Nominatim rate limit

    with open(RESEARCH, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=fields)
        writer.writeheader()
        writer.writerows(rows)

    print(f"\nGeocoded {done}, failed {fail}. Wrote {RESEARCH}")


if __name__ == "__main__":
    main()
