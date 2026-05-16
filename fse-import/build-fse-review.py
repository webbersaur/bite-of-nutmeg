#!/usr/bin/env python3
"""
Build a review spreadsheet from FSE.xlsx (Ledge Light Health District
food-service permit list) for abiteofnutmeg.com town expansion.

- Separates each establishment by CT town (parses messy addresses, folds
  village names into their town, treats Mystic as its own town).
- Classifies each as KEEP (restaurant / bar / cafe / deli / bakery) or
  REMOVE (grocery, gas, fraternal club, school, institutional, etc.).
- Flags low-confidence calls with a star for Chris to review.

Output: FSE-review.xlsx (written next to this script)

Run from anywhere:  python3 fse-import/build-fse-review.py
"""

import os
import re
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

HERE = os.path.dirname(os.path.abspath(__file__))
SRC = os.path.join(HERE, "FSE.xlsx")
OUT = os.path.join(HERE, "FSE-review.xlsx")

# ---------------------------------------------------------------- towns
TOWNS = ["New London", "Waterford", "East Lyme", "Old Lyme", "Lyme", "Ledyard",
         "North Stonington", "Stonington", "Groton", "Montville", "Norwich"]

VILLAGE = {
    "pawcatuck": "Stonington", "niantic": "East Lyme", "gales ferry": "Groton",
    "noank": "Groton", "quaker hill": "Waterford", "uncasville": "Montville",
    "oakdale": "Montville", "hadlyme": "Lyme", "flanders": "East Lyme",
    "poquetanuck": "Ledyard",
}


def parse_town(addr):
    """Resolve a CT town from a free-form address string."""
    if not addr:
        return "?"
    a = " ".join(str(addr).split())
    low = a.lower()
    # Mystic is its own town (catches "Mystic" and "Old Mystic")
    if re.search(r"\bmystic\b", low):
        return "Mystic"
    # explicit (Town) parenthetical
    m = re.search(r"\(([^)]+)\)", a)
    if m:
        p = m.group(1).strip().lower()
        for t in TOWNS:
            if t.lower() == p:
                return t
    # full town name anywhere (longest first so "East Lyme" beats "Lyme")
    for t in sorted(TOWNS, key=len, reverse=True):
        if re.search(r"\b" + re.escape(t.lower()) + r"\b", low):
            return t
    # village -> town
    for v, t in VILLAGE.items():
        if re.search(r"\b" + v + r"\b", low):
            return t
    return "?"


# ----------------------------------------------------------- classifier
# Each REMOVE rule: (regex, reason). First match wins.
REMOVE_RULES = [
    (r"\b(elementary|middle school|high school|magnet|campus)\b", "school"),
    (r"\bschool\b", "school"),
    (r"\b(college|university)\b|conncoll|mitchell college", "college"),
    (r"chartwells|compass group|sodexo", "college dining contractor"),
    (r"\b(church|congregation|parish|baptist|methodist|episcopal|"
     r"unitarian|ministry|temple emanu|our lady|christ the king|"
     r"holy ghost|retreat)\b", "church / religious"),
    (r"\b(nursing|rehab|snf|assisted living|health ?care|hospital|"
     r"congregate|masonicare|institute|infirmary)\b|senior living|"
     r"senior center|acute care|\bmanor\b|crossroads place|"
     r"\bbeechwood\b", "senior / care facility"),
    (r"^camp\b|\b(daycare|head ?start|child development|learning center|"
     r"learning academy|pre-?school|montessori|early childhood)\b|"
     r"\bkidds\b|cherished children|precious memories", "childcare / camp"),
    (r"\bfire (department|company)\b", "fire department"),
    (r"\b(shell|citgo|sunoco|valero|mobil|gulf|exxon|getty|bp)\b|"
     r"food & fuel|fast fuel|henny penny|town stop", "gas station"),
    (r"\b(market|marketplace|supermarket|grocery|bodega|wholesale|aldi|"
     r"costco|walmart|target|shoprite|shop ?rite|provisions|depot)\b|"
     r"mini ?mart|food ?mart|quick mart|\bmart\b|big y|stop ?and ?shop|"
     r"stop ?& ?shop|dollar general|dollar tree|dollar & up|\bbestway\b|"
     r"cumberland farms|cash & carry|general store|country market|"
     r"food co-?op|farmstand|farm stand|\borchards?\b|7-?eleven|alltown|"
     r"\bstore\b|pay rite|amg retail|best ?way", "grocery / market"),
    (r"\b(cvs|walgreens|pharmacy)\b|rite aid|ollie's|bargain outlet|"
     r"\bflorist\b|greenery|news ?stand|smoke and more|spice and tea|"
     r"\bconfections?\b|chocolates?\b|edible arrangements|pinspiration|"
     r"pepper palace", "retail / non-dining"),
    (r"pfizer|electric boat|\bdominion\b|millstone|b605|wet dock|"
     r"self service area|davis-standard", "corporate cafeteria"),
    (r"american legion|\belks\b|\bvfw\b|knights of columbus|\blodge\b|"
     r"fleet reserve|\bsociety\b|(country|yacht|beach|workingm|"
     r"dramatic|sportsm|german|vet|athletic|social)[\w']*\s+club|"
     r"\bveterans?\b|submarine vet", "fraternal / club"),
    (r"community meal|meal center|breakfast program|\bshelter\b|"
     r"salvation army|community center|recreation center|"
     r"neighborhood center|alliance for living|soup kitchen|tvcca",
     "community / social services"),
    (r"theatre|theater|playhouse|cinemas?|arts center|nature center|"
     r"\bmuseum\b|art cinemas", "arts / entertainment venue"),
    (r"\bbowling\b", "bowling alley"),
    (r"\bcatering\b|\bcaterer\b", "caterer"),
    (r"\bnutrition\b|atto di fede|blast of energy|\benergy\b", "nutrition club"),
    (r"personal training", "personal training"),
    (r"\b(concessions?|boosters?)\b|little league|"
     r"youth league|youth football|snack shack", "concession / booster"),
    (r"\bcafeteria\b", "institutional cafeteria"),
    (r"hampton inn|hilton|hyatt|marriott|holiday inn|sleep inn|super 8|"
     r"best western|residence inn|spring hill suites|rodeway|"
     r"spark by hilton|regency inn|\bmotel\b|inn & suites|inn and suites|"
     r"hotel & spa", "hotel / lodging"),
]

# names a REMOVE rule would wrongly catch, or REMOVE calls no rule catches.
# value: (decision, category_or_reason, uncertain)
OVERRIDE = {
    # genuine restaurants a REMOVE rule would otherwise flag
    "Grille 92 at Fairview":                         ("KEEP", "American", False),
    "Slipper Shell Galley @ Niantic Bay Yacht Club": ("KEEP", "Seafood", False),
    "K&S Custom Catering LLC dba Norm's Diner":      ("KEEP", "Diner", False),
    "Dog Watch Mystic & Dog Watch Catering":         ("KEEP", "Bar & Kitchen", False),
    "Pequot Golf & Grille":                          ("KEEP", "American", True),
    "Michael's Dairy":                               ("KEEP", "Ice Cream", False),
    "Boro Bodega and Sccopery":                      ("KEEP", "Ice Cream", True),
    "Smoothie King":                                 ("KEEP", "Cafe / Smoothie", True),
    "Puerto Lima - Farm To Bowl":                    ("KEEP", "Restaurant", True),
    "Flanders Fish Market & Restaurant":             ("KEEP", "Seafood", False),
    "Moe's Southwest Grill - Store #3554":           ("KEEP", "Mexican / Latin", False),
    "JJ Sunset LLC d/b/a McDonalds at WalMart #18676": ("KEEP", "Fast Food", False),
    "NV Bakery and Market":                          ("KEEP", "Cafe & Bakery", False),
    "Andy's Deli & Market":                          ("KEEP", "Deli / Sandwich", False),
    "Haylon's Market/Deke's Bagels":                 ("KEEP", "Cafe & Bakery", True),
    "Mystic Market":                                 ("REMOVE", "gourmet market (has cafe?)", True),
    "Mystic Market East":                            ("REMOVE", "gourmet market (has cafe?)", True),
    # non-dining establishments no REMOVE rule catches
    "Academy Point at Mystic":         ("REMOVE", "senior living", False),
    "Crescent Point at Niantic":       ("REMOVE", "senior living", False),
    "StoneRidge":                      ("REMOVE", "senior living", False),
    "Complete Care at Groton Regency LLC": ("REMOVE", "senior / care facility", False),
    "Harbour Village North":           ("REMOVE", "senior living", False),
    "Altruism Acute Care & Evaluation":("REMOVE", "health care", False),
    "Fairview":                        ("REMOVE", "senior community", False),
    "Club Demonstration Services":     ("REMOVE", "in-store demo vendor", False),
    "Black Hall Club, Inc.":           ("REMOVE", "private club", False),
    "Thames Club":                     ("REMOVE", "private club", False),
    "Wadawanuck Club":                 ("REMOVE", "private club", False),
    "Tri Town Foods":                  ("REMOVE", "grocery / market", False),
    "East Lyme Noble, LLC":            ("REMOVE", "gas station", True),
    "Best Way":                        ("REMOVE", "convenience store", False),
    "Best Way of North Stonington 2":  ("REMOVE", "convenience store", False),
    "Mystic Dark Room":                ("REMOVE", "retail (photography)", False),
    "Barkin' Barley":                  ("REMOVE", "retail (pet bakery)", False),
    "Honey Bee Farms":                 ("REMOVE", "farm stand", False),
    "Sankow's Beaver Brook Farm, LLC": ("REMOVE", "farm", False),
    "Mystic River Chocolate":          ("REMOVE", "retail (chocolate)", False),
    "Capizzanos Oils & Vinegars":      ("REMOVE", "retail (specialty food)", False),
    "The Spice Palette":               ("REMOVE", "retail (specialty food)", True),
    "The Spice Club":                  ("REMOVE", "retail (specialty food)", True),
    "Tiger Lily Tea":                  ("REMOVE", "retail (tea)", True),
    "Thames River Greenery":           ("REMOVE", "retail (garden)", False),
    "The Tin Peddler":                 ("REMOVE", "retail (gift shop)", False),
    "The Carousel Shop":               ("REMOVE", "retail (gift shop)", True),
    "The Ditty Bag LLC":               ("REMOVE", "retail (gift shop)", True),
    "Lamplighter Trading Company, LLC":("REMOVE", "retail", True),
    "Gumdrops & Lollipops":            ("REMOVE", "retail (candy)", True),
    "ATY Confections":                 ("REMOVE", "retail (confections)", True),
    "Koastal Krunch":                  ("REMOVE", "retail / packaged food", True),
    "Lake of Isles":                   ("REMOVE", "golf resort", True),
    "Clubhouse at Elmridge":           ("REMOVE", "golf clubhouse", True),
    "New London Sports Complex":       ("REMOVE", "sports venue", True),
    "Great Brook Sports":              ("REMOVE", "sports venue", True),
    "Ella T. Grasso":                  ("REMOVE", "technical school", False),
    "DJ's Campus Kitchen":             ("REMOVE", "college dining", True),
    "Boost Bowls":                     ("REMOVE", "nutrition club", False),
    "Fit Fam":                         ("REMOVE", "nutrition club", True),
    "Prepped LLC":                     ("REMOVE", "meal prep", True),
    "GG at Home":                      ("REMOVE", "meal delivery", True),
    "Gather 360":                      ("REMOVE", "catering / events", True),
    "Clyde's Cider Mill":              ("REMOVE", "farm / attraction", True),
    "Birdseye, LLC":                   ("REMOVE", "unclear", True),
    "Steve Tyler":                     ("REMOVE", "individual permit", True),
    "Fadia Al-Hasan":                  ("REMOVE", "individual permit", True),
    "Sarwat Qamar":                    ("REMOVE", "individual permit", True),
}

# KEEP category detection -- first match wins.
CATEGORY_RULES = [
    ("Pizza",          r"pizz|apizza|\bcrust\b|brick ?oven|wood ?fired|"
                       r"s'?barro"),
    ("Bar",            r"tavern|\bpub\b|alehouse|ale house|gastropub|brew|"
                       r"\bbar\b|cocktail|\blounge\b|cabaret"),
    ("Winery",         r"winery|vineyards?|tasting room"),
    ("Cafe & Bakery",  r"\bcafe\b|caf\xe9|coffee|espresso|bakery|bakeshop|"
                       r"\bbake\b|donut|doughnut|bagel|pastry|crepe|cr\xeape|"
                       r"\bwaffle\b|\bsweet\b|crumb|muffin|starbucks|dunkin|"
                       r"panera"),
    ("Ice Cream",      r"creamery|ice cream|\bfroyo\b|frozen yogurt|"
                       r"water ice|italian ice|custard|gelato|\bdairy\b"),
    ("Seafood",        r"seafood|\blobster\b|\boyster\b|\bclam\b|\bcrab\b|"
                       r"chowder|\bshanty\b|sea swirl"),
    ("Chinese",        r"\bchina\b|chinese|\bwok\b|dumpling|peking|chopstick|"
                       r"hunan|szechuan"),
    ("Japanese",       r"sushi|hibachi|japanese|steakhouse|ramen|\bkoto\b"),
    ("Thai",           r"\bthai\b"),
    ("Indian",         r"\bindia\b|indian|tandoori|masala"),
    ("Mexican / Latin",r"mexican|taqueria|taquerio|\btaco\b|burrito|cantina|"
                       r"pupusa|pollos|chapina|caribbean"),
    ("Italian",        r"italian|trattoria|osteria|\bpasta\b|piadina"),
    ("BBQ",            r"\bbbq\b|barbecue|smokehouse"),
    ("Deli / Sandwich",r"\bdeli\b|grinder|sandwich|\bsubs?\b|\bwraps?\b"),
    ("Diner",          r"\bdiner\b"),
    ("American",       r"grille?|kitchen|restaurant|eatery|bistro|burger|"
                       r"\bwings?\b|\bchicken\b|rotisserie|smashburger"),
    ("Fast Food",      r"mcdonald|burger king|\bwendy|popeyes|subway|"
                       r"applebee|denny'?s|chili'?s|texas roadhouse|"
                       r"olive garden|longhorn|moe'?s|five guys|"
                       r"kentucky fried|taco bell|\bkfc\b"),
]

KEEP_RE = re.compile(
    "|".join(r for _, r in CATEGORY_RULES), re.I)


def guess_category(name):
    for cat, rx in CATEGORY_RULES:
        if re.search(rx, name, re.I):
            return cat
    return ""


def clean(text):
    """Collapse all whitespace (incl. non-breaking spaces) to single spaces."""
    return " ".join(str(text or "").split())


def classify(name):
    """Return (decision, category_or_reason, uncertain_bool)."""
    n = clean(name)
    if n in OVERRIDE:
        return OVERRIDE[n]
    for rx, reason in REMOVE_RULES:
        if re.search(rx, n, re.I):
            return ("REMOVE", reason, False)
    cat = guess_category(n)
    if cat:
        return ("KEEP", cat, False)
    # no signal either way -- keep, but flag for review
    return ("KEEP", "Restaurant?", True)


# --------------------------------------------------------------- build
def main():
    wb = openpyxl.load_workbook(SRC, read_only=True, data_only=True)
    src = wb[wb.sheetnames[0]]
    rows = list(src.iter_rows(values_only=True))[1:]

    records = []
    for r in rows:
        name, addr, email, permit = r[0], r[1], r[2], r[3]
        town = parse_town(addr)
        decision, cat, uncertain = classify(name)
        records.append({
            "town": town, "decision": decision, "category": cat,
            "uncertain": uncertain, "name": clean(name), "addr": clean(addr),
            "permit": permit, "email": clean(email),
        })

    # sort: town, kept-first, name
    records.sort(key=lambda x: (x["town"], 0 if x["decision"] == "KEEP" else 1,
                                x["name"].lower()))

    # ---- write workbook
    out = openpyxl.Workbook()
    ws = out.active
    ws.title = "FSE Review"

    navy = "1E3A6E"
    gold = "F0B323"
    grey = "E2E2E2"
    hdr_font = Font(bold=True, color="FFFFFF", size=11)
    hdr_fill = PatternFill("solid", fgColor=navy)
    title_font = Font(bold=True, size=13, color=navy)
    sub_font = Font(italic=True, size=10, color="555555")
    remove_fill = PatternFill("solid", fgColor=grey)
    star_fill = PatternFill("solid", fgColor=gold)
    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # summary block
    towns = sorted({r["town"] for r in records})
    ws["A1"] = "A Bite of Nutmeg — FSE.xlsx Town/Dining Review"
    ws["A1"].font = title_font
    ws["A2"] = ("KEEP = restaurant/bar/cafe/deli/bakery   •   "
                "REMOVE = grocery/gas/club/school/institutional   •   "
                "★ = low confidence, please review")
    ws["A2"].font = sub_font

    ws["A4"] = "Town"
    ws["B4"] = "Keep"
    ws["C4"] = "Remove"
    ws["D4"] = "Total"
    for c in "ABCD":
        cell = ws[f"{c}4"]
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = Alignment(horizontal="center")
    row = 5
    tot_k = tot_r = 0
    for t in towns:
        recs = [r for r in records if r["town"] == t]
        k = sum(1 for r in recs if r["decision"] == "KEEP")
        rm = len(recs) - k
        tot_k += k
        tot_r += rm
        ws[f"A{row}"] = t
        ws[f"B{row}"] = k
        ws[f"C{row}"] = rm
        ws[f"D{row}"] = len(recs)
        for c in "ABCD":
            ws[f"{c}{row}"].border = border
        row += 1
    ws[f"A{row}"] = "TOTAL"
    ws[f"B{row}"] = tot_k
    ws[f"C{row}"] = tot_r
    ws[f"D{row}"] = tot_k + tot_r
    for c in "ABCD":
        ws[f"{c}{row}"].font = Font(bold=True)
        ws[f"{c}{row}"].border = border
    row += 2

    # main table header
    headers = ["Town", "Keep", "Category / Reason", "★",
               "Facility Name", "Address", "Permit #", "Email"]
    hdr_row = row
    for i, h in enumerate(headers, 1):
        cell = ws.cell(row=hdr_row, column=i, value=h)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = Alignment(horizontal="center")
    row += 1

    for r in records:
        vals = [r["town"], r["decision"], r["category"],
                "★" if r["uncertain"] else "", r["name"], r["addr"],
                r["permit"], r["email"]]
        for i, v in enumerate(vals, 1):
            cell = ws.cell(row=row, column=i, value=v)
            cell.border = border
            if r["decision"] == "REMOVE":
                cell.fill = remove_fill
            if i == 4 and r["uncertain"]:
                cell.fill = star_fill
                cell.alignment = Alignment(horizontal="center")
        row += 1

    # column widths / freeze / autofilter
    widths = [16, 9, 22, 4, 36, 46, 9, 30]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[chr(64 + i)].width = w
    ws.freeze_panes = f"A{hdr_row + 1}"
    ws.auto_filter.ref = f"A{hdr_row}:H{row - 1}"

    out.save(OUT)
    print(f"Processed {len(records)} records.")
    unresolved = [r for r in records if r["town"] == "?"]
    print(f"Unresolved towns: {len(unresolved)}")
    for r in unresolved:
        print("   ", r["name"], "|", r["addr"])
    print(f"KEEP: {tot_k}   REMOVE: {tot_r}   "
          f"uncertain: {sum(1 for r in records if r['uncertain'])}")
    print(f"Wrote {OUT}")


if __name__ == "__main__":
    main()
